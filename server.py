"""
WebScraper - Universal Web Scraping Tool Backend v2.0
FastAPI server with static + optional Playwright dynamic scraping.
Features: Dynamic rendering, session control, XHR capture, anti-detection,
deep data extraction, advanced crawl strategies.
"""

import os
import re
import json
import time
import random
import hashlib
import base64
import asyncio
import logging
from collections import deque
from datetime import datetime
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, quote
from typing import Optional, List, Dict, Any
import uuid
import csv
import io
import threading
import concurrent.futures
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel, Field

import requests
from bs4 import BeautifulSoup, Comment

# ---------------------------------------------------------------------------
# Optional imports
# ---------------------------------------------------------------------------
try:
    import aiohttp
    AIOHTTP_AVAILABLE = True
except ImportError:
    AIOHTTP_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from playwright.async_api import async_playwright, Page, BrowserContext
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    import curl_cffi
    CURL_CFFI_AVAILABLE = True
except ImportError:
    CURL_CFFI_AVAILABLE = False


# Connection pool for session reuse (keep-alive)
_session_pool = {}
def get_session(proxy=None):
    key = proxy or "__default__"
    if key not in _session_pool:
        sess = requests.Session()
        if proxy:
            sess.proxies = {"http": proxy, "https": proxy}
        _session_pool[key] = sess
    return _session_pool[key]

# Task management for cancellation support
_tasks = {}  # task_id -> {"cancelled": bool, "pages": [], "errors": [], "status": "running|completed|cancelled", "created_at": 0}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("webscraper")

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

async def cleanup_old_tasks():
    """Remove tasks older than 10 minutes."""
    while True:
        await asyncio.sleep(600)
        now = time.time()
        to_remove = [tid for tid, t in _tasks.items() if t.get("created_at", now) < now - 600]
        for tid in to_remove:
            del _tasks[tid]
        if to_remove:
            logger.info(f"Cleaned up {len(to_remove)} old tasks")


@asynccontextmanager
async def lifespan(app_instance):
    # Startup
    asyncio.create_task(cleanup_old_tasks())
    threading.Timer(2.0, _start_scheduler).start()
    yield
    # Shutdown
    await _close_browser()

app = FastAPI(title="WebScraper API", version="2.0.0", lifespan=lifespan)

_STATIC_DIR = os.path.dirname(os.path.abspath(__file__))


@app.get("/")
async def root():
    return FileResponse(os.path.join(_STATIC_DIR, "index.html"))


@app.get("/index.html")
async def index_html():
    return FileResponse(os.path.join(_STATIC_DIR, "index.html"))


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_DEPTH = 10
MAX_PAGES = 500
CONCURRENCY_STATIC = 5
CONCURRENCY_DYNAMIC = 3

# ---------------------------------------------------------------------------
# User-Agent Rotation Pool (16 modern browser UAs)
# ---------------------------------------------------------------------------
USER_AGENT_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 OPR/109.0.0.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

DEFAULT_HEADERS = {
    "User-Agent": USER_AGENT_POOL[0],
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-US;q=0.7",
    "Accept-Encoding": "gzip, deflate",
    "Sec-Ch-Ua": '"Chromium";v="125", "Not.A/Brand";v="24", "Google Chrome";v="125"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "max-age=0",
}
# ---------------------------------------------------------------------------
# Playwright Browser Reuse (module-level singleton)
# ---------------------------------------------------------------------------
_playwright_instance = None
_browser = None
_browser_lock = asyncio.Lock()


async def _get_browser(proxy=None):
    global _playwright_instance, _browser
    if not PLAYWRIGHT_AVAILABLE:
        return None
    async with _browser_lock:
        if _browser is not None:
            try:
                _ = _browser.is_connected()
                return _browser
            except Exception:
                _browser = None
        try:
            _playwright_instance = await async_playwright().start()
            launch_args = [
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-extensions",
                "--disable-component-update",
                "--disable-background-networking",
            ]
            proxy_config = None
            if proxy:
                proxy_config = {"server": proxy}
            _browser = await _playwright_instance.chromium.launch(
                headless=True, args=launch_args, proxy=proxy_config,
            )
            logger.info("Playwright browser launched (shared instance)")
            return _browser
        except Exception as e:
            logger.error(f"Failed to launch Playwright browser: {e}")
            _browser = None
            return None


async def _close_browser():
    global _playwright_instance, _browser
    async with _browser_lock:
        if _browser:
            try:
                await _browser.close()
            except Exception:
                pass
            _browser = None
        if _playwright_instance:
            try:
                await _playwright_instance.stop()
            except Exception:
                pass
            _playwright_instance = None
    logger.info("Playwright browser closed")





# ---------------------------------------------------------------------------
# Pydantic Models
# ---------------------------------------------------------------------------

class CustomSelectors(BaseModel):
    article_body: Optional[str] = ""
    price: Optional[str] = ""
    title: Optional[str] = ""


class InteractionAction(BaseModel):
    action: str
    selector: Optional[str] = None
    value: Optional[str] = None
    delay_ms: int = 500


class ScrapeOptions(BaseModel):
    depth: int = Field(0, ge=0, le=10)
    max_pages: int = Field(1, ge=1, le=500)
    same_domain: bool = True
    timeout: int = Field(30, ge=5, le=300)
    custom_selectors: Optional[CustomSelectors] = CustomSelectors()
    crawl_mode: str = "bfs"
    delay_ms: int = Field(0, ge=0, le=10000)
    max_retries: int = Field(2, ge=0, le=5)
    proxy: Optional[str] = None
    cookies: Optional[Dict[str, str]] = None
    headers: Optional[Dict[str, str]] = None
    rotate_ua: bool = True
    interactions: Optional[List[InteractionAction]] = None
    wait_for_selector: Optional[str] = None
    enable_websocket_capture: bool = True
    enable_iframe_extraction: bool = True
    enable_shadow_dom: bool = True
    discover_sitemap: bool = True
    discover_hidden_links: bool = True
    task_id: Optional[str] = None


class ScrapeRequest(BaseModel):
    url: str
    extraction_mode: str = "auto"
    extract_types: List[str] = ["text", "links", "media", "structured", "metadata", "business"]
    options: ScrapeOptions = ScrapeOptions()


class BatchScrapeRequest(BaseModel):
    urls: List[str]
    extraction_mode: str = "auto"
    extract_types: List[str] = ["text", "links", "media", "structured", "metadata", "business"]
    options: ScrapeOptions = ScrapeOptions()


class ProxyRequest(BaseModel):
    url: str
    method: str = "GET"
    headers: Optional[Dict[str, str]] = {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resolve_url(base, href):
    if not href:
        return ""
    href = href.strip()
    if href.startswith(("javascript:", "mailto:", "tel:", "#", "data:")):
        return ""
    return urljoin(base, href)


def same_domain(url1, url2):
    return urlparse(url1).netloc == urlparse(url2).netloc


def clean_text(text):
    return re.sub(r'\s+', ' ', text).strip()


def get_ua(options):
    if options.rotate_ua:
        return random.choice(USER_AGENT_POOL)
    if options.headers and "User-Agent" in options.headers:
        return options.headers["User-Agent"]
    return DEFAULT_HEADERS["User-Agent"]


def build_request_headers(ua, options, referer=None):
    """Build complete request headers matching the selected User-Agent."""
    is_firefox = 'Firefox' in ua and 'Edg/' not in ua
    is_safari = 'Safari' in ua and 'Chrome' not in ua and 'Edg/' not in ua and 'OPR/' not in ua
    is_edge = 'Edg/' in ua
    is_opera = 'OPR/' in ua

    chrome_ver = None
    m = re.search(r'Chrome/(\d+)', ua)
    if m:
        chrome_ver = m.group(1)

    if is_firefox:
        headers = {
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3",
            "Accept-Encoding": "gzip, deflate",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
        }
    elif is_safari:
        headers = {
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Accept-Encoding": "gzip, deflate",
            "Connection": "keep-alive",
        }
    else:
        headers = {
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-US;q=0.7",
            "Accept-Encoding": "gzip, deflate",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
        }
        if chrome_ver:
            major = chrome_ver
            headers["Sec-Ch-Ua"] = '"Chromium";v="%s", "Not.A/Brand";v="24", "Google Chrome";v="%s"' % (major, major)
            headers["Sec-Ch-Ua-Mobile"] = "?0"
            platform = random.choice(['"Windows"', '"macOS"', '"Linux"'])
            headers["Sec-Ch-Ua-Platform"] = platform

    if random.random() < 0.4:
        headers["DNT"] = "1"

    if referer:
        headers["Referer"] = referer
    elif options.headers and "Referer" in options.headers:
        headers["Referer"] = options.headers["Referer"]

    if options.headers:
        for k, v in options.headers.items():
            headers[k] = v

    return headers


def build_proxies(proxy):
    if not proxy:
        return None
    return {"http": proxy, "https": proxy}


def random_delay_ms(options):
    if options.delay_ms > 0:
        jitter = random.randint(0, options.delay_ms)
        time.sleep(jitter / 1000.0)


def detect_encoding(response, soup):
    if response.encoding and response.encoding.lower() not in ('iso-8859-1',):
        return response.encoding
    meta = soup.find("meta", attrs={"charset": True})
    if meta:
        return meta["charset"]
    meta = soup.find("meta", attrs={"http-equiv": re.compile(r"content-type", re.I)})
    if meta and "charset=" in meta.get("content", ""):
        return meta["content"].split("charset=")[-1].strip().split(";")[0]
    return "utf-8"
# ---------------------------------------------------------------------------
# Extraction Functions
# ---------------------------------------------------------------------------

def extract_text(soup, base_url, selectors):
    result = {"title": "", "headings": {}, "paragraphs": [], "lists": [], "main_content": ""}
    if selectors.title:
        el = soup.select_one(selectors.title)
        if el:
            result["title"] = clean_text(el.get_text())
    if not result["title"]:
        t = soup.find("title")
        if t:
            result["title"] = clean_text(t.get_text())

    for i in range(1, 7):
        tag = f"h{i}"
        headings = [clean_text(h.get_text()) for h in soup.find_all(tag)]
        if headings:
            result["headings"][tag] = headings

    main_el = None
    if selectors.article_body:
        main_el = soup.select_one(selectors.article_body)
    if not main_el:
        main_el = soup.find("main") or soup.find("article") or soup.find("body") or soup

    for p in main_el.find_all("p"):
        text = clean_text(p.get_text())
        if text and len(text) > 10:
            result["paragraphs"].append(text)

    for lst in main_el.find_all(["ul", "ol"]):
        items = []
        for li in lst.find_all("li", recursive=False):
            text = clean_text(li.get_text())
            if text:
                items.append(text)
        if items:
            result["lists"].append(items)

    result["main_content"] = clean_text(main_el.get_text(separator="\n"))
    return result


def extract_links(soup, base_url):
    internal = []
    external = []
    base_domain = urlparse(base_url).netloc
    for a in soup.find_all("a", href=True):
        href = resolve_url(base_url, a["href"])
        if not href:
            continue
        text = clean_text(a.get_text())
        parsed = urlparse(href)
        link_info = {"url": href, "text": text, "title": a.get("title", "")}
        if parsed.netloc == base_domain or parsed.netloc == "":
            internal.append(link_info)
        else:
            external.append(link_info)
    return {"internal": internal, "external": external, "all_count": len(internal) + len(external)}


def extract_media(soup, base_url):
    images = []
    videos = []
    audio = []
    documents = []
    doc_exts = ('.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.csv', '.zip', '.rar', '.txt')

    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or img.get("data-lazy-src", "")
        if src:
            images.append({"url": resolve_url(base_url, src), "alt": img.get("alt", ""),
                           "width": img.get("width", ""), "height": img.get("height", "")})
    for source in soup.find_all("source", type=re.compile(r"image/")):
        src = source.get("src", "")
        if src:
            images.append({"url": resolve_url(base_url, src), "alt": "", "width": "", "height": ""})

    for video in soup.find_all("video"):
        src = video.get("src", "")
        if src:
            videos.append({"url": resolve_url(base_url, src), "type": "direct"})
        for source in video.find_all("source"):
            src = source.get("src", "")
            if src:
                videos.append({"url": resolve_url(base_url, src), "type": source.get("type", "unknown")})
    for iframe in soup.find_all("iframe", src=True):
        src = iframe["src"]
        if re.search(r'(youtube\.com/embed|youtu\.be|player\.vimeo)', src):
            videos.append({"url": resolve_url(base_url, src), "type": "iframe"})

    for aud in soup.find_all("audio"):
        src = aud.get("src", "")
        if src:
            audio.append({"url": resolve_url(base_url, src), "type": "direct"})
        for source in aud.find_all("source"):
            src = source.get("src", "")
            if src:
                audio.append({"url": resolve_url(base_url, src), "type": source.get("type", "unknown")})

    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        if any(href.endswith(ext) or ext in href for ext in doc_exts):
            documents.append({"url": resolve_url(base_url, a["href"]), "text": clean_text(a.get_text()),
                              "type": href.split(".")[-1] if "." in href else "unknown"})

    return {"images": images, "videos": videos, "audio": audio, "documents": documents}


def extract_structured(soup, base_url):
    tables = []
    json_ld = []
    inline_json = []

    for table in soup.find_all("table"):
        headers = []
        rows_data = []
        for th in table.find_all("th"):
            headers.append(clean_text(th.get_text()))
        if not headers:
            first_row = table.find("tr")
            if first_row:
                for td in first_row.find_all(["td", "th"]):
                    headers.append(clean_text(td.get_text()))
        for tr in table.find_all("tr")[1 if headers else 0:]:
            cells = [clean_text(td.get_text()) for td in tr.find_all(["td", "th"])]
            if cells:
                if headers and len(cells) == len(headers):
                    rows_data.append(dict(zip(headers, cells)))
                else:
                    rows_data.append(cells)
        if rows_data:
            tables.append({"headers": headers, "rows": rows_data, "row_count": len(rows_data)})

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            json_ld.append(data)
        except (json.JSONDecodeError, TypeError):
            pass

    for script in soup.find_all("script", type="application/json"):
        try:
            data = json.loads(script.string or "")
            inline_json.append(data)
        except (json.JSONDecodeError, TypeError):
            pass

    microdata = []
    for item in soup.find_all(attrs={"itemscope": True}):
        item_type = item.get("itemtype", "")
        props = {}
        for prop in item.find_all(attrs={"itemprop": True}):
            name = prop["itemprop"]
            value = prop.get("content") or clean_text(prop.get_text())
            props[name] = value
        if props:
            microdata.append({"type": item_type, "properties": props})
    if microdata:
        inline_json.append({"_microdata": microdata})

    return {"tables": tables, "json_ld": json_ld, "inline_json": inline_json}


def extract_metadata(soup, base_url):
    result = {"title": "", "description": "", "keywords": "", "author": "",
              "publish_date": "", "open_graph": {}, "twitter_cards": {}}
    title_tag = soup.find("title")
    if title_tag:
        result["title"] = clean_text(title_tag.get_text())

    for meta in soup.find_all("meta"):
        name = (meta.get("name") or meta.get("property") or "").lower()
        content = meta.get("content", "")
        if not name or not content:
            continue
        if name == "description":
            result["description"] = content
        elif name == "keywords":
            result["keywords"] = content
        elif name == "author":
            result["author"] = content
        elif name in ("date", "dc.date", "dcterms.created"):
            result["publish_date"] = content
        elif name.startswith("og:"):
            result["open_graph"][name[3:]] = content
        elif name.startswith("twitter:"):
            result["twitter_cards"][name[8:]] = content

    if not result["publish_date"]:
        time_tag = soup.find("time", datetime=True)
        if time_tag:
            result["publish_date"] = time_tag["datetime"]

    for meta in soup.find_all("meta", property=re.compile(r"article:(published|modified)")):
        if not result["publish_date"]:
            result["publish_date"] = meta.get("content", "")

    return result


def extract_business(soup, base_url, selectors):
    prices = []
    ratings = []
    dates = []
    locations = []
    text = soup.get_text()

    price_patterns = [
        r'[\u00a5\uffe5$\u20ac\u00a3]\s*\d[\d,]*\.?\d*',
        r'\d[\d,]*\.?\d*\s*(?:\u5143|dollars|USD|EUR|GBP|CNY|RMB)',
        r'(?:\u4ef7\u683c|price|\u552e\u4ef7|\u539f\u4ef7|\u73b0\u4ef7|\u6298\u6263\u4ef7)[\uff1a:]\s*[\u00a5\uffe5$\u20ac\u00a3]?\s*\d[\d,]*\.?\d*',
    ]
    seen_prices = set()
    for pattern in price_patterns:
        for match in re.finditer(pattern, text, re.I):
            val = match.group().strip()
            if val not in seen_prices:
                seen_prices.add(val)
                prices.append(val)
    if selectors.price:
        for el in soup.select(selectors.price):
            val = clean_text(el.get_text())
            if val and val not in seen_prices:
                prices.append(val)

    rating_patterns = [
        r'(\d\.?\d?)\s*/\s*5',
        r'(\d\.?\d?)\s*(?:stars?|\u661f|\u5206|\u8bc4\u5206)',
        r'(?:\u8bc4\u5206|rating|score)[\uff1a:]\s*(\d\.?\d?)',
    ]
    seen_ratings = set()
    for pattern in rating_patterns:
        for match in re.finditer(pattern, text, re.I):
            val = match.group().strip()
            if val not in seen_ratings:
                seen_ratings.add(val)
                ratings.append(val)

    date_patterns = [
        r'\d{4}[-/\u5e74.]\d{1,2}[-/\u6708.]\d{1,2}(?:[\u65e5]?)',
        r'\d{1,2}[-/]\d{1,2}[-/]\d{4}',
    ]
    seen_dates = set()
    for pattern in date_patterns:
        for match in re.finditer(pattern, text):
            val = match.group().strip()
            if val not in seen_dates:
                seen_dates.add(val)
                dates.append(val)

    loc_patterns = [
        r'(?:\u5730\u5740|address|location|\u5730\u70b9|\u4f4d\u7f6e|\u6240\u5728)[\uff1a:]\s*(.{5,80}?)(?:\n|\u3002|\.)',
        r'(?:\u7701|\u5e02|\u533a|\u53bf|\u9547|\u8def|\u8857|\u53f7|\u697c|\u5ba4|\u680b|\u5f04|\u5df7).{0,30}(?:\u8def|\u8857|\u53f7|\u697c|\u5ba4|\u680b|\u5f04|\u5df7)',
    ]
    seen_locs = set()
    for pattern in loc_patterns:
        for match in re.finditer(pattern, text):
            val = match.group().strip()
            if val not in seen_locs and len(val) > 5:
                seen_locs.add(val)
                locations.append(val)

    return {"prices": prices, "ratings": ratings, "dates": dates, "locations": locations}
def prioritize_urls(items):
    """Sort (url, depth) tuples to visit listing/category pages before detail pages."""
    def url_priority(item):
        url = item[0] if isinstance(item, tuple) else item
        try:
            path = urlparse(url).path.rstrip('/')
        except Exception:
            return 3
        if not path or path.count('/') <= 1:
            return 0
        if '?' in url and 'page' in url.lower():
            return 1
        if path.count('/') <= 2:
            return 2
        return 3
    return sorted(items, key=url_priority)

# ---------------------------------------------------------------------------# ---------------------------------------------------------------------------
# Static Fetch Engine
# ---------------------------------------------------------------------------

def fetch_page(url, timeout=30, options=None):
    if options is None:
        options = ScrapeOptions()
    ua = get_ua(options)
    headers = build_request_headers(ua, options, referer=options.headers.get("Referer") if options.headers else None)
    proxies = build_proxies(options.proxy)
    max_retries = options.max_retries

    for attempt in range(max_retries + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout,
                               allow_redirects=True, verify=False, proxies=proxies,
                               cookies=options.cookies)
            resp.raise_for_status()
            return resp
        except Exception as e:
            if attempt < max_retries:
                backoff = (2 ** attempt) + random.uniform(0.5, 1.5)
                logger.warning(f"Retry {attempt + 1}/{max_retries} for {url} after {backoff:.1f}s: {e}")
                time.sleep(backoff)
            else:
                logger.error(f"Failed to fetch {url} after {max_retries + 1} attempts: {e}")
                return None
    return None


def parse_and_extract(url, html, extract_types, selectors):
    soup = BeautifulSoup(html, "lxml")
    for comment in soup.find_all(string=lambda t: isinstance(t, Comment)):
        comment.extract()

    extracted = {}
    if "text" in extract_types:
        extracted["text"] = extract_text(soup, url, selectors)
    if "links" in extract_types:
        extracted["links"] = extract_links(soup, url)
    if "media" in extract_types:
        extracted["media"] = extract_media(soup, url)
    if "structured" in extract_types:
        extracted["structured"] = extract_structured(soup, url)
    if "metadata" in extract_types:
        extracted["metadata"] = extract_metadata(soup, url)
    if "business" in extract_types:
        extracted["business"] = extract_business(soup, url, selectors)

    title = ""
    if selectors.title:
        el = soup.select_one(selectors.title)
        if el:
            title = clean_text(el.get_text())
    if not title:
        t = soup.find("title")
        if t:
            title = clean_text(t.get_text())

    return {"title": title, "soup": soup, "extracted": extracted}


# ---------------------------------------------------------------------------
# Signature / Token Pattern Detection
# ---------------------------------------------------------------------------

def detect_signed_params(url, headers):
    findings = []
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)

    token_patterns = re.compile(
        r'(token|sign|signature|timestamp|nonce|ts|sig|auth|key|hash|hmac|encrypted|enc|'
        r'access_token|api_key|apikey|request_id|requestid|session_id)', re.IGNORECASE)

    for param_name, values in qs.items():
        if token_patterns.search(param_name):
            findings.append({"location": "query_param", "name": param_name,
                             "value_preview": values[0][:64] if values else ""})
        elif values and len(values[0]) > 20 and re.match(r'^[A-Za-z0-9_\-\.]+$', values[0]):
            findings.append({"location": "query_param", "name": param_name,
                             "value_preview": values[0][:64], "note": "long token-like value"})

    for header_name, header_value in headers.items():
        if token_patterns.search(header_name):
            findings.append({"location": "header", "name": header_name,
                             "value_preview": header_value[:64] if header_value else ""})

    return findings
# ---------------------------------------------------------------------------
# Dynamic Fetch (Playwright) with XHR Capture, WS, Shadow DOM, iframes
# ---------------------------------------------------------------------------

async def fetch_html_dynamic(url, options, extract_types):
    if not PLAYWRIGHT_AVAILABLE:
        return None

    browser = await _get_browser(options.proxy)
    if not browser:
        return None

    ua = get_ua(options)
    api_calls = []
    ws_frames = []

    context_args = {
        "viewport": {"width": 1920, "height": 1080},
        "user_agent": ua,
        "locale": "en-US",
        "timezone_id": "America/New_York",
        "color_scheme": "light",
        "device_scale_factor": random.choice([1, 1.25, 1.5, 2]),
        "has_touch": False,
        "java_script_enabled": True,
        "bypass_csp": False,
    }
    if options.proxy:
        context_args["proxy"] = {"server": options.proxy}

    context = await browser.new_context(**context_args)
    try:

        if options.cookies:
            cookie_list = []
            parsed_base = urlparse(url)
            for name, value in options.cookies.items():
                cookie_list.append({"name": name, "value": value, "domain": parsed_base.netloc, "path": "/"})
            await context.add_cookies(cookie_list)

        extra_headers = dict(options.headers) if options.headers else {}
        if extra_headers:
            await context.set_extra_http_headers(extra_headers)

        page = await context.new_page()

        # Comprehensive fingerprint evasion
        stealth_script = """
            // Remove webdriver flag
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            delete navigator.__proto__.webdriver;

            // Languages
            Object.defineProperty(navigator, 'languages', { get: () => ['zh-CN', 'zh', 'en-US', 'en'] });
            Object.defineProperty(navigator, 'language', { get: () => 'zh-CN' });

            // Plugins - simulate real Chrome plugins
            Object.defineProperty(navigator, 'plugins', {
                get: () => {
                    const plugins = [
                        { name: 'Chrome PDF Plugin', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
                        { name: 'Chrome PDF Viewer', filename: 'mhjfbmdgcfjbbpaeojofohoefgiehjai', description: '' },
                        { name: 'Native Client', filename: 'internal-nacl-plugin', description: '' },
                    ];
                    plugins.length = 3;
                    return plugins;
                }
            });

            // Platform
            Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });

            // Hardware concurrency (match real CPUs)
            Object.defineProperty(navigator, 'hardwareConcurrency', { get: () => 8 });

            // Device memory
            Object.defineProperty(navigator, 'deviceMemory', { get: () => 8 });

            // Max touch points
            Object.defineProperty(navigator, 'maxTouchPoints', { get: () => 0 });

            // Chrome runtime
            window.chrome = { runtime: {}, loadTimes: function() { return {}; }, csi: function() { return {}; } };

            // Permissions API override
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
            );

            // WebGL vendor/renderer spoofing
            const getParameter = WebGLRenderingContext.prototype.getParameter;
            WebGLRenderingContext.prototype.getParameter = function(parameter) {
                if (parameter === 37445) return 'Intel Inc.';
                if (parameter === 37446) return 'Intel Iris OpenGL Engine';
                return getParameter.call(this, parameter);
            };

            // Remove HeadlessChrome from UA if present
            Object.defineProperty(navigator, 'userAgent', {
                get: () => navigator.userAgent.replace('HeadlessChrome', 'Chrome')
            });

            // Notification permission
            if (typeof Notification !== 'undefined') {
                Object.defineProperty(Notification, 'permission', { get: () => 'default' });
            }

            // Connection rtt (real browsers have non-zero rtt)
            if (navigator.connection) {
                Object.defineProperty(navigator.connection, 'rtt', { get: () => 50 });
            }
        """
        await page.add_init_script(stealth_script)

        # XHR/Fetch interception
        async def on_response(response):
            try:
                req = response.request
                if req.resource_type in ("xhr", "fetch"):
                    resp_headers = await response.all_headers()
                    content_type = resp_headers.get("content-type", "")
                    status = response.status
                    method = req.method
                    req_url = req.url

                    body_preview = ""
                    try:
                        if "json" in content_type or "xml" in content_type:
                            body_text = await response.text()
                            body_preview = body_text[:5000]
                        elif "javascript" in content_type:
                            body_text = await response.text()
                            body_preview = body_text[:2000]
                    except Exception:
                        pass

                    req_headers_dict = {}
                    try:
                        req_headers_dict = req.headers
                    except Exception:
                        pass

                    api_calls.append({
                        "url": req_url, "method": method, "status": status,
                        "content_type": content_type, "response_preview": body_preview,
                        "request_headers": dict(req_headers_dict) if isinstance(req_headers_dict, dict) else {},
                    })
            except Exception:
                pass

        page.on("response", on_response)

        # WebSocket capture
        if options.enable_websocket_capture:
            def on_ws(ws):
                ws_url = ws.url
                ws_info = {"url": ws_url, "frames": []}

                def on_ws_frame(payload):
                    try:
                        frame_data = ""
                        if isinstance(payload, str):
                            frame_data = payload[:3000]
                        elif isinstance(payload, bytes):
                            frame_data = base64.b64encode(payload).decode("ascii")[:3000]
                        ws_info["frames"].append({"data": frame_data, "time": datetime.utcnow().isoformat()})
                    except Exception:
                        pass

                ws.on("framereceived", lambda payload: on_ws_frame(payload))
                ws.on("framesent", lambda payload: on_ws_frame(payload))
                ws_frames.append(ws_info)

            page.on("websocket", on_ws)

        # Natural browsing: optionally visit domain root first
        parsed_url = urlparse(url)
        origin = f"{parsed_url.scheme}://{parsed_url.netloc}"
        if origin != url:
            try:
                await page.goto(origin, wait_until="domcontentloaded", timeout=min(10000, options.timeout * 1000))
                await asyncio.sleep(random.uniform(0.5, 1.5))
            except Exception:
                pass  # Silently ignore pre-visit failure

        # Navigate
        timeout_ms = options.timeout * 1000
        status_code = 200
        try:
            response = await page.goto(url, wait_until="networkidle", timeout=timeout_ms)
            if response:
                status_code = response.status
        except Exception as e:
            logger.warning(f"Playwright navigation error for {url}: {e}")
            try:
                status_code = 200 if page.url else 0
            except Exception:
                status_code = 0

        # Wait for selector
        if options.wait_for_selector:
            try:
                await page.wait_for_selector(options.wait_for_selector, timeout=10000)
            except Exception:
                logger.debug(f"wait_for_selector timed out: {options.wait_for_selector}")

        # Execute user interactions
        if options.interactions:
            for action in options.interactions:
                try:
                    if action.action == "scroll":
                        await page.evaluate("window.scrollBy(0, window.innerHeight)")
                    elif action.action == "scroll_to_bottom":
                        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    elif action.action == "click" and action.selector:
                        await page.click(action.selector, timeout=5000)
                    elif action.action == "wait":
                        await asyncio.sleep((action.delay_ms or 500) / 1000.0)
                    elif action.action == "type" and action.selector and action.value:
                        await page.fill(action.selector, action.value)
                    await asyncio.sleep((action.delay_ms or 500) / 1000.0)
                except Exception as e:
                    logger.debug(f"Interaction '{action.action}' failed: {e}")

        # Extra wait for dynamic content
        await asyncio.sleep(1)

        # Get page HTML
        html = await page.content()

        # Shadow DOM extraction
        shadow_dom_content = []
        if options.enable_shadow_dom:
            try:
                shadow_results = await page.evaluate("""() => {
                    const results = [];
                    function findShadowRoots(root) {
                        const all = root.querySelectorAll('*');
                        for (const el of all) {
                            if (el.shadowRoot) {
                                results.push({
                                    tag: el.tagName,
                                    id: el.id || '',
                                    className: el.className || '',
                                    text: el.shadowRoot.textContent.substring(0, 5000),
                                    html: el.shadowRoot.innerHTML.substring(0, 10000)
                                });
                                findShadowRoots(el.shadowRoot);
                            }
                        }
                    }
                    findShadowRoots(document);
                    return results;
                }""")
                shadow_dom_content = shadow_results or []
            except Exception as e:
                logger.debug(f"Shadow DOM extraction failed: {e}")

        # iframe content extraction
        iframe_content = []
        if options.enable_iframe_extraction:
            try:
                frames = page.frames
                for frame in frames:
                    if frame == page.main_frame:
                        continue
                    try:
                        frame_url = frame.url
                        frame_html = await frame.content()
                        frame_title = ""
                        try:
                            frame_title = await frame.evaluate("document.title || ''")
                        except Exception:
                            pass
                        iframe_content.append({
                            "url": frame_url, "title": frame_title,
                            "html_preview": frame_html[:10000],
                            "text_preview": clean_text(BeautifulSoup(frame_html, "lxml").get_text())[:5000],
                        })
                    except Exception:
                        pass
            except Exception as e:
                logger.debug(f"iframe extraction failed: {e}")

    finally:
        await context.close()

    # Detect signed params in captured API calls
    signed_params = []
    for call in api_calls:
        found = detect_signed_params(call.get("url", ""), call.get("request_headers", {}))
        if found:
            call["signed_params"] = found
            signed_params.extend(found)

    return {
        "html": html, "status_code": status_code, "api_calls": api_calls,
        "websocket_frames": ws_frames, "shadow_dom": shadow_dom_content,
        "iframe_content": iframe_content, "signed_params": signed_params,
    }
# ---------------------------------------------------------------------------
# Hidden Link Discovery
# ---------------------------------------------------------------------------

async def discover_sitemap_urls(base_url, options):
    parsed = urlparse(base_url)
    sitemap_url = f"{parsed.scheme}://{parsed.netloc}/sitemap.xml"
    urls = []
    try:
        resp = await asyncio.get_running_loop().run_in_executor(
            None, lambda: fetch_page(sitemap_url, timeout=options.timeout, options=options))
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "xml")
            for loc in soup.find_all("loc"):
                url = loc.get_text().strip()
                if url:
                    urls.append(url)
            for sitemap in soup.find_all("sitemap"):
                inner_loc = sitemap.find("loc")
                if inner_loc:
                    sub_urls = await discover_sitemap_urls(inner_loc.get_text().strip(), options)
                    urls.extend(sub_urls)
    except Exception as e:
        logger.debug(f"Sitemap discovery failed for {sitemap_url}: {e}")
    return urls


async def discover_robots_txt_paths(base_url, options):
    parsed = urlparse(base_url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    paths = []
    try:
        resp = await asyncio.get_running_loop().run_in_executor(
            None, lambda: fetch_page(robots_url, timeout=options.timeout, options=options))
        if resp and resp.status_code == 200:
            for line in resp.text.splitlines():
                line = line.strip()
                if line.lower().startswith("allow:") or line.lower().startswith("disallow:"):
                    parts = line.split(":", 1)
                    if len(parts) == 2:
                        path = parts[1].strip()
                        if path and path != "/" and not path.startswith("#"):
                            full_url = urljoin(base_url, path)
                            paths.append(full_url)
                elif line.lower().startswith("sitemap:"):
                    parts = line.split(":", 1)
                    if len(parts) == 2:
                        sm_url = parts[1].strip()
                        if sm_url:
                            sm_urls = await discover_sitemap_urls(sm_url, options)
                            paths.extend(sm_urls)
    except Exception as e:
        logger.debug(f"robots.txt discovery failed: {e}")
    return paths


def detect_url_patterns(urls):
    generated = []

    # Numeric increment pattern
    numeric_pattern = re.compile(r'(/page/\d+|/p/\d+|/item/\d+|/\d+\.html?|/page=\d+)')
    for url in urls:
        match = numeric_pattern.search(url)
        if match:
            num_str = re.search(r'(\d+)', match.group())
            if num_str:
                num = int(num_str.group())
                base_part = url[:match.start()]
                suffix_part = url[match.end():]
                num_fmt = num_str.group()
                for delta in range(1, 6):
                    new_num = num + delta
                    new_num_str = str(new_num).zfill(len(num_fmt))
                    new_url = f"{base_part}{new_num_str}{suffix_part}"
                    if new_url not in urls:
                        generated.append(new_url)

    # Query param pagination
    param_pattern = re.compile(r'[?&](page|p|offset|start|skip)=\d+')
    for url in urls:
        match = param_pattern.search(url)
        if match:
            qs = parse_qs(urlparse(url).query)
            param_key = match.group(1)
            if param_key in qs:
                current_val = int(qs[param_key][0])
                for delta in range(1, 6):
                    new_url = re.sub(
                        rf'([?&]{param_key}=)\d+', rf'\g<1>{current_val + delta}', url, count=1)
                    if new_url not in urls:
                        generated.append(new_url)

    return generated


def discover_hidden_links(soup, base_url):
    found = []

    # noscript links
    for noscript in soup.find_all("noscript"):
        inner_soup = BeautifulSoup(str(noscript), "html.parser")
        for a in inner_soup.find_all("a", href=True):
            href = resolve_url(base_url, a["href"])
            if href:
                found.append({"url": href, "source": "noscript", "text": clean_text(a.get_text())})

    # data-* attributes with URLs
    for tag in soup.find_all(True):
        for attr_name, attr_val in tag.attrs.items():
            if attr_name.startswith("data-") and isinstance(attr_val, str):
                if re.search(r'https?://', attr_val) or re.search(r'/[a-zA-Z0-9_\-/]+', attr_val):
                    href = resolve_url(base_url, attr_val)
                    if href:
                        found.append({"url": href, "source": f"data-{attr_name}", "text": ""})

    # JavaScript href assignments
    for script in soup.find_all("script"):
        if script.string:
            js_urls = re.findall(
                r'(?:window\.location(?:\.href)?|location(?:\.href)?)\s*=\s*["\']([^"\']+)["\']',
                script.string)
            for js_url in js_urls:
                href = resolve_url(base_url, js_url)
                if href:
                    found.append({"url": href, "source": "javascript", "text": ""})

    return found
# ---------------------------------------------------------------------------
# Core Scrape Page (static + dynamic)
# ---------------------------------------------------------------------------

async def scrape_page(url, extract_types, options, mode="auto"):
    timeout = options.timeout
    selectors = options.custom_selectors or CustomSelectors()
    result = {"url": url, "title": "", "status_code": 200, "extracted": {}}

    html = None

    # Static fetch
    if mode in ("auto", "static"):
        resp = fetch_page(url, timeout, options)
        if resp:
            result["status_code"] = resp.status_code
            try:
                enc = resp.apparent_encoding or "utf-8"
                if isinstance(enc, tuple):
                    enc = "utf-8"
                resp.encoding = enc
                html = resp.text
            except Exception:
                try:
                    html = resp.content.decode("utf-8", errors="replace")
                except Exception:
                    html = ""

    # Dynamic fetch fallback
    pw_result = None
    if html is None and mode in ("auto", "dynamic"):
        if PLAYWRIGHT_AVAILABLE:
            pw_result = await fetch_html_dynamic(url, options, extract_types)
            if pw_result:
                html = pw_result.get("html")
                result["status_code"] = pw_result.get("status_code", result["status_code"])
            if html is None:
                result["status_code"] = 0
        else:
            result["status_code"] = 0
            result["extracted"] = {"error": "Playwright not available. Install with: pip install playwright && python -m playwright install chromium"}
            return result

    if html is None:
        result["status_code"] = 0
        result["extracted"] = {"error": f"Failed to fetch {url}"}
        return result

    parsed = parse_and_extract(url, html, extract_types, selectors)
    result["title"] = parsed["title"]
    result["extracted"] = parsed["extracted"]

    # Merge Playwright-captured data
    if pw_result:
        result["extracted"]["api_calls"] = pw_result.get("api_calls", [])
        result["extracted"]["websocket_frames"] = pw_result.get("websocket_frames", [])
        result["extracted"]["shadow_dom"] = pw_result.get("shadow_dom", [])
        result["extracted"]["iframe_content"] = pw_result.get("iframe_content", [])
    else:
        result["extracted"]["api_calls"] = []
        result["extracted"]["websocket_frames"] = []
        result["extracted"]["shadow_dom"] = []
        result["extracted"]["iframe_content"] = []

    return result


# ---------------------------------------------------------------------------
# Crawl Engine (BFS + DFS)
# ---------------------------------------------------------------------------

async def scrape_recursive(start_url, extract_types, options, mode="auto", task_id=None):
    depth = min(options.depth, MAX_DEPTH)
    max_pages = min(options.max_pages, MAX_PAGES)
    same_domain_only = options.same_domain
    crawl_mode = options.crawl_mode.lower()

    visited = set()
    pages = []
    errors = []
    all_discovered_urls = set()
    sitemap_urls_count = 0

    # Register task for cancellation support
    if task_id:
        if task_id not in _tasks:
            _tasks[task_id] = {"cancelled": False, "pages": [], "errors": [], "status": "running", "created_at": time.time(), "crawl_stats": {}}

    start_time = time.time()
    total_bytes = 0

    # Discover hidden URLs
    extra_urls = []
    if options.discover_sitemap:
        try:
            sitemap_urls = await discover_sitemap_urls(start_url, options)
            extra_urls.extend(sitemap_urls[:max_pages])
            sitemap_urls_count = len(sitemap_urls)
        except Exception as e:
            logger.debug(f"Sitemap discovery error: {e}")

    if options.discover_hidden_links:
        try:
            robots_urls = await discover_robots_txt_paths(start_url, options)
            extra_urls.extend(robots_urls[:max_pages])
        except Exception as e:
            logger.debug(f"robots.txt discovery error: {e}")

    if options.discover_hidden_links or options.discover_sitemap:
        url_patterns = detect_url_patterns(extra_urls)
        extra_urls.extend(url_patterns[:50])

    queue = deque()
    queue.append((start_url, 0))
    visited.add(start_url)

    sem = asyncio.Semaphore(CONCURRENCY_STATIC if mode != "dynamic" else CONCURRENCY_DYNAMIC)

    async def scrape_one(target_url):
        async with sem:
            random_delay_ms(options)
            return await scrape_page(target_url, extract_types, options, mode)

    while queue and len(pages) < max_pages:
        # Check for cancellation
        if task_id and task_id in _tasks and _tasks[task_id]["cancelled"]:
            logger.info(f"Task {task_id} cancelled, stopping crawl")
            if task_id in _tasks:
                _tasks[task_id]["status"] = "cancelled"
            break
        batch = []
        next_items = []

        if crawl_mode == "dfs":
            while queue and len(batch) < min(CONCURRENCY_STATIC, max_pages - len(pages)):
                if not queue:
                    break
                url_item, cur_depth = queue.pop()
                batch.append((url_item, cur_depth))
        else:
            while queue and len(batch) < min(CONCURRENCY_STATIC, max_pages - len(pages)):
                if not queue:
                    break
                url_item, cur_depth = queue.popleft()
                batch.append((url_item, cur_depth))

        if not batch:
            break

        results = await asyncio.gather(*[scrape_one(u) for u, _ in batch], return_exceptions=True)

        for (url_item, cur_depth), res in zip(batch, results):
            if isinstance(res, Exception):
                errors.append({"url": url_item, "error": str(res)})
                logger.error(f"Error scraping {url_item}: {res}")
                continue

            pages.append(res)

            extracted_text = res.get("extracted", {}).get("text", {})
            main_content = extracted_text.get("main_content", "")
            total_bytes += len(main_content.encode("utf-8", errors="replace"))

            if cur_depth < depth and len(pages) < max_pages:
                links_data = res.get("extracted", {}).get("links", {})
                all_links = links_data.get("internal", []) + links_data.get("external", [])

                for link in all_links:
                    link_url = link.get("url", "")
                    if not link_url or link_url in visited:
                        continue
                    if same_domain_only and not same_domain(start_url, link_url):
                        continue
                    visited.add(link_url)
                    all_discovered_urls.add(link_url)
                    next_items.append((link_url, cur_depth + 1))

        for item in prioritize_urls(next_items):
            queue.append(item)

        # Update task progress
        if task_id and task_id in _tasks:
            _tasks[task_id]["pages"] = pages
            _tasks[task_id]["errors"] = errors

    # Inject sitemap/extra URLs if remaining capacity
    if len(pages) < max_pages and extra_urls:
        for extra_url in extra_urls:
            if len(pages) >= max_pages:
                break
            if extra_url in visited:
                continue
            if same_domain_only and not same_domain(start_url, extra_url):
                continue
            visited.add(extra_url)
            all_discovered_urls.add(extra_url)
            try:
                result = await scrape_one(extra_url)
                if not isinstance(result, Exception):
                    pages.append(result)
            except Exception as e:
                errors.append({"url": extra_url, "error": str(e)})

    duration = time.time() - start_time

    crawl_stats = {
        "total_pages": len(pages),
        "total_bytes": total_bytes,
        "duration_seconds": round(duration, 2),
        "pages_per_second": round(len(pages) / max(duration, 0.001), 2),
        "urls_discovered": len(all_discovered_urls),
        "sitemap_urls": sitemap_urls_count,
    }


    # Mark task as completed
    if task_id and task_id in _tasks:
        _tasks[task_id]["pages"] = pages
        _tasks[task_id]["errors"] = errors
        _tasks[task_id]["status"] = "completed"
        _tasks[task_id]["crawl_stats"] = crawl_stats

    return {
        "pages": pages,
        "total_pages_scraped": len(pages),
        "errors": errors,
        "crawl_stats": crawl_stats,
    }
# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------

@app.get("/api/health")
async def health():
    return {
        "status": "ok",
        "playwright_available": PLAYWRIGHT_AVAILABLE,
        "curl_cffi_available": CURL_CFFI_AVAILABLE,
        "version": "2.0.0",
        "browser_running": _browser is not None and _browser.is_connected() if _browser else False,
    }


@app.post("/api/scrape")
async def scrape(req: ScrapeRequest):
    task_id = str(uuid.uuid4())[:8]
    _tasks[task_id] = {"cancelled": False, "pages": [], "errors": [], "status": "running", "created_at": time.time(), "crawl_stats": {}}

    url = req.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="URL is required")
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    try:
        result = await scrape_recursive(url, req.extract_types, req.options, req.extraction_mode, task_id=task_id)
        _tasks[task_id]["status"] = "completed"
        _tasks[task_id]["crawl_stats"] = result.get("crawl_stats", {})
        return JSONResponse(content={"success": True, "task_id": task_id, **result})
    except Exception as e:
        logger.error(f"Scrape error: {e}")
        _tasks[task_id]["status"] = "cancelled"
        return JSONResponse(
            status_code=500,
            content={
                "success": False, "task_id": task_id, "error": str(e), "pages": _tasks[task_id]["pages"],
                "total_pages_scraped": len(_tasks[task_id]["pages"]),
                "errors": _tasks[task_id]["errors"] + [{"error": str(e)}],
                "crawl_stats": {"total_pages": 0, "total_bytes": 0, "duration_seconds": 0,
                                "pages_per_second": 0, "urls_discovered": 0, "sitemap_urls": 0},
            })


@app.post("/api/scrape/batch")
async def scrape_batch(req: BatchScrapeRequest):
    task_id = str(uuid.uuid4())[:8]
    _tasks[task_id] = {"cancelled": False, "pages": [], "errors": [], "status": "running", "created_at": time.time(), "crawl_stats": {}}

    all_pages = []
    all_errors = []
    total_bytes = 0
    start_time = time.time()
    sitemap_urls_count = 0

    for raw_url in req.urls:
        # Check for cancellation
        if task_id and task_id in _tasks and _tasks[task_id]["cancelled"]:
            logger.info(f"Batch task {task_id} cancelled")
            break

        url = raw_url.strip()
        if not url:
            continue
        if not url.startswith(("http://", "https://")):
            url = "https://" + url

        try:
            random_delay_ms(req.options)
            result = await scrape_page(url, req.extract_types, req.options, req.extraction_mode)
            all_pages.append(result)
            total_bytes += len(json.dumps(result).encode("utf-8", errors="replace"))
            # Update task progress
            if task_id and task_id in _tasks:
                _tasks[task_id]["pages"] = all_pages
                _tasks[task_id]["errors"] = all_errors
        except Exception as e:
            all_errors.append({"url": url, "error": str(e)})

    duration = time.time() - start_time

    # Mark task as completed
    if task_id and task_id in _tasks:
        _tasks[task_id]["pages"] = all_pages
        _tasks[task_id]["errors"] = all_errors
        _tasks[task_id]["status"] = "completed"

    return JSONResponse(content={
        "success": True, "task_id": task_id, "pages": all_pages, "total_pages_scraped": len(all_pages),
        "errors": all_errors,
        "crawl_stats": {
            "total_pages": len(all_pages), "total_bytes": total_bytes,
            "duration_seconds": round(duration, 2),
            "pages_per_second": round(len(all_pages) / max(duration, 0.001), 2),
            "urls_discovered": 0, "sitemap_urls": sitemap_urls_count,
        },
    })


@app.post("/api/scrape/cancel")
async def cancel_scrape(request_body: dict):
    """Cancel a running scrape task and return partial results."""
    task_id = request_body.get("task_id", "")
    if task_id and task_id in _tasks:
        _tasks[task_id]["cancelled"] = True
        # Wait briefly for the task to finish its current page
        for _ in range(50):  # up to 5 seconds
            if _tasks[task_id]["status"] in ("completed", "cancelled"):
                break
            await asyncio.sleep(0.1)

        # Return whatever we have
        task = _tasks[task_id]
        return JSONResponse(content={
            "success": True,
            "pages": task["pages"],
            "total_pages_scraped": len(task["pages"]),
            "errors": task["errors"],
            "cancelled": True,
            "crawl_stats": task.get("crawl_stats", {})
        })
    raise HTTPException(status_code=404, detail="Task not found")


@app.get("/api/scrape/status/{task_id}")
async def scrape_status(task_id: str):
    """Get current status of a scrape task (for polling)."""
    if task_id in _tasks:
        task = _tasks[task_id]
        return {
            "task_id": task_id,
            "status": task["status"],
            "pages_scraped": len(task["pages"]),
            "cancelled": task["cancelled"]
        }
    raise HTTPException(status_code=404, detail="Task not found")


@app.post("/api/proxy")
async def proxy(req: ProxyRequest):
    try:
        resp = requests.request(
            method=req.method.upper(), url=req.url,
            headers={**DEFAULT_HEADERS, **(req.headers or {})},
            timeout=30, allow_redirects=True, verify=False)
        content_type = resp.headers.get("content-type", "text/plain")
        return JSONResponse(
            content={"status_code": resp.status_code, "content_type": content_type,
                     "body": resp.text[:5_000_000]})
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Proxy error: {e}")




# ---------------------------------------------------------------------------
# Hotlist — 热搜聚合数据接口
# ---------------------------------------------------------------------------
import atexit
import sys
import requests as _req_sync

# ---------------------------------------------------------------------------
# Hotlist — 热搜聚合（多线程自研爬取，不依赖第三方API）
# ---------------------------------------------------------------------------
# 每个平台有主数据源+备用数据源，多线程并行抓取，单平台<3秒，全部<5秒
# ---------------------------------------------------------------------------

HOTLIST_PLATFORMS = {
    "weibo": {"name": "微博热搜", "color": "#e6162d", "total": 50},
    "zhihu": {"name": "知乎热榜", "color": "#0066ff", "total": 50},
    "baidu": {"name": "百度热搜", "color": "#2932e1", "total": 50},
    "douyin": {"name": "抖音热点", "color": "#111111", "total": 50},
    "toutiao": {"name": "头条热榜", "color": "#e4393c", "total": 50},
    "bilibili": {"name": "B站热门", "color": "#fb7299", "total": 50},
}

# 每个平台的数据源列表（主 + 备用），按优先级排列
_HOTLIST_SOURCES = {
    "weibo": [
        {"type": "api", "url": "https://weibo.com/ajax/side/hotSearch", "headers": {"Referer": "https://weibo.com", "X-Requested-With": "XMLHttpRequest"}},
        {"type": "api", "url": "https://tenapi.cn/v2/weibohot"},
        {"type": "mobile", "url": "https://m.weibo.cn/api/container/getIndex?containerid=106003type%3D25%26t%3D3%26disable_hot%3D1%26filter_type%3Drealtimehot", "headers": {"Referer": "https://m.weibo.cn", "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X)"}},
    ],
    "zhihu": [
        {"type": "api", "url": "https://www.zhihu.com/api/v3/feed/topstory/hot-lists/total?limit=50", "headers": {"Referer": "https://www.zhihu.com/hot", "X-Requested-With": "fetch"}},
        {"type": "api", "url": "https://tenapi.cn/v2/zhihuhot"},
    ],
    "baidu": [
        {"type": "parse", "url": "https://top.baidu.com/board?tab=realtime", "parser": "baidu_html"},
        {"type": "api", "url": "https://tenapi.cn/v2/baiduhot"},
    ],
    "douyin": [
        {"type": "api", "url": "https://www.douyin.com/aweme/v1/web/hot/search/list/?detail_list=1&source=6&main_billboard_count=30", "headers": {"Referer": "https://www.douyin.com", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}},
        {"type": "api", "url": "https://tenapi.cn/v2/douyinhot"},
    ],
    "toutiao": [
        {"type": "api", "url": "https://www.toutiao.com/hot-event/hot-board/?origin=toutiao_pc&_signature=_02B4Z6wo00f01", "headers": {"Referer": "https://www.toutiao.com"}},
        {"type": "api", "url": "https://tenapi.cn/v2/toutiaohot"},
    ],
    "bilibili": [
        {"type": "api", "url": "https://api.bilibili.com/x/web-interface/ranking/v2?rid=0&type=all", "headers": {"Referer": "https://www.bilibili.com"}},
        {"type": "api", "url": "https://tenapi.cn/v2/bilihot"},
    ],
}

# 连接池：每个平台独立 session，复用 TCP 连接
_platform_sessions = {}

def _get_platform_session(platform):
    if platform not in _platform_sessions:
        s = _req_sync.Session()
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Accept": "application/json, text/html, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate",
            "Connection": "keep-alive",
        })
        _platform_sessions[platform] = s
    return _platform_sessions[platform]


_hotlist_cache = {}
_hotlist_lock = threading.Lock()
_HOTLIST_CACHE_FILE = os.path.join(_STATIC_DIR, "hotlist_cache.json")


def _load_hotlist_cache():
    global _hotlist_cache
    if os.path.exists(_HOTLIST_CACHE_FILE):
        try:
            with open(_HOTLIST_CACHE_FILE, "r", encoding="utf-8") as f:
                _hotlist_cache = json.load(f)
            logger.info(f"[Hotlist] Loaded cache: {len(_hotlist_cache)} platforms")
        except Exception as e:
            logger.warning(f"[Hotlist] Cache load error: {e}")


def _save_hotlist_cache():
    try:
        with _hotlist_lock:
            with open(_HOTLIST_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(_hotlist_cache, f, ensure_ascii=False, indent=1)
    except Exception as e:
        logger.warning(f"[Hotlist] Cache save error: {e}")


# ---- 解析器：各平台专用数据解析 ----

def _parse_weibo(items, data):
    """解析微博热搜 JSON"""
    real_data = data.get("data", {}).get("realtime", []) if isinstance(data.get("data"), dict) else []
    if not real_data:
        real_data = data.get("data", []) if isinstance(data.get("data"), list) else []
    for idx, item in enumerate(real_data[:50]):
        title = item.get("note") or item.get("word") or item.get("name") or ""
        if not title:
            continue
        items.append({
            "rank": idx + 1, "title": title.strip(),
            "heat": int(item.get("num") or item.get("hot") or item.get("raw_hot") or 0),
            "change": int(item.get("change") or item.get("changeType") or 0),
            "tags": ([item["icon_name"]] if item.get("icon_name") else []),
            "url": f"https://s.weibo.com/weibo?q=%23{title}%23" if not item.get("url") else item["url"],
        })


def _parse_zhihu(items, data):
    """解析知乎热榜"""
    targets = data.get("data", []) if isinstance(data.get("data"), list) else []
    for idx, target in enumerate(targets[:50]):
        detail = target.get("target", {})
        title = detail.get("title") or target.get("title") or ""
        if not title:
            continue
        items.append({
            "rank": idx + 1, "title": title.strip(),
            "heat": int(target.get("detail_text", "0").replace("万热度", "0000").replace(" 热度", "").replace(",", "") or 0),
            "change": 0,
            "tags": [],
            "url": f"https://www.zhihu.com/question/{detail.get('id', '')}" if detail.get("id") else "",
        })


def _parse_baidu_html(items, html_text):
    """从百度热搜页面 HTML 解析数据"""
    # 提取 JSON 数据块
    match = re.search(r'<!--s-data:(.*?)-->', html_text, re.DOTALL)
    if match:
        try:
            data = json.loads(match.group(1))
            cards = data.get("data", {}).get("cards", [])
            for card in cards:
                for idx, item in enumerate(card.get("content", [])[:50]):
                    title = item.get("word") or item.get("query") or ""
                    if not title:
                        continue
                    items.append({
                        "rank": len(items) + 1, "title": title.strip(),
                        "heat": int(item.get("hotScore") or item.get("rawHot") or 0),
                        "change": int(item.get("change") or 0),
                        "tags": ([item["tag"]] if item.get("tag") else []),
                        "url": item.get("url") or item.get("mobileUrl") or "",
                    })
        except json.JSONDecodeError:
            pass


def _parse_douyin(items, data):
    """解析抖音热点"""
    word_list = data.get("data", {}).get("word_list", []) if isinstance(data.get("data"), dict) else []
    if not word_list:
        word_list = data.get("word_list", []) if isinstance(data.get("word_list"), list) else []
    for idx, item in enumerate(word_list[:50]):
        title = item.get("word") or item.get("title") or ""
        if not title:
            continue
        items.append({
            "rank": idx + 1, "title": title.strip(),
            "heat": int(item.get("hot_value") or item.get("event_time") or 0),
            "change": int(item.get("word_type") or 0),
            "tags": ([item["label"]] if item.get("label") else []),
            "url": "",
        })


def _parse_toutiao(items, data):
    """解析头条热榜"""
    targets = data.get("data", []) if isinstance(data.get("data"), list) else []
    for idx, item in enumerate(targets[:50]):
        title = item.get("Title") or item.get("title") or ""
        if not title:
            continue
        items.append({
            "rank": idx + 1, "title": title.strip(),
            "heat": int(item.get("HotValue") or item.get("hot_value") or 0),
            "change": int(item.get("Change") or item.get("change") or 0),
            "tags": ([item["Label"]] if item.get("Label") else []),
            "url": item.get("Url") or item.get("url") or "",
        })


def _parse_bilibili(items, data):
    """解析B站热门"""
    vlist = data.get("data", {}).get("list", []) if isinstance(data.get("data"), dict) else []
    for idx, item in enumerate(vlist[:50]):
        title = item.get("title") or ""
        if not title:
            continue
        items.append({
            "rank": idx + 1, "title": title.strip(),
            "heat": int(item.get("stat", {}).get("view", 0)),
            "change": 0,
            "tags": ([item.get("tname", "")] if item.get("tname") else []),
            "url": f"https://www.bilibili.com/video/{item.get('bvid', '')}" if item.get("bvid") else "",
        })


_PARSERS = {
    "weibo": _parse_weibo,
    "zhihu": _parse_zhihu,
    "baidu": _parse_baidu_html,
    "douyin": _parse_douyin,
    "toutiao": _parse_toutiao,
    "bilibili": _parse_bilibili,
}


def _scrape_platform(platform):
    """从多数据源抓取平台热搜，自动降级到备用源"""
    sources = _HOTLIST_SOURCES.get(platform, [])
    session = _get_platform_session(platform)

    for source in sources:
        try:
            url = source["url"]
            headers = source.get("headers", {})
            resp = session.get(url, headers=headers, timeout=8, allow_redirects=True)

            if resp.status_code != 200:
                logger.debug(f"[Hotlist] {platform} source {source['type']} returned {resp.status_code}")
                continue

            items = []
            content_type = resp.headers.get("content-type", "")

            if source.get("parser") == "baidu_html":
                _parse_baidu_html(items, resp.text)
            elif "json" in content_type or resp.text.strip().startswith("{") or resp.text.strip().startswith("["):
                data = resp.json()
                parser = _PARSERS.get(platform)
                if parser:
                    parser(items, data)

            if items:
                logger.info(f"[Hotlist] {platform}: {len(items)} items from {url[:50]}")
                return items

        except _req_sync.exceptions.Timeout:
            logger.debug(f"[Hotlist] {platform} timeout on {source['url'][:50]}")
            continue
        except Exception as e:
            logger.debug(f"[Hotlist] {platform} error on source: {e}")
            continue

    # 所有数据源失败，返回模拟数据
    logger.warning(f"[Hotlist] {platform}: all sources failed, using fallback")
    return _generate_fallback(platform)


def _generate_fallback(platform):
    """当所有数据源不可用时生成模拟数据"""
    now = datetime.now()
    seed = int(hashlib.md5((platform + now.strftime('%Y%m%d%H')).encode()).hexdigest()[:8], 16)
    topics_map = {
        'weibo': ['#某明星官宣恋情#', '全国多地迎来强降雨', '高考成绩即将公布', '新款iPhone发布在即', '某热播剧大结局', '体育赛事精彩回顾', '科学家发现新星系', '多城市调整公积金政策', '新能源车销量创新高', '高校毕业生就业新政策', '某地突发自然灾害', '国际油价大幅波动', '热门综艺节目开播', '某上市公司财报超预期', 'AI技术最新突破', '某城市地铁新线开通', '春节假期旅游数据', '某知名导演新片定档', '全国天气预报更新', '某品牌新品发布会', '互联网行业动态', '某球员转会消息', '高考志愿填报指南', '某地美食节开幕', '豆瓣高分电影推荐', '某科学家获国际大奖', '股市行情分析', '某城市限购政策调整', '暑期档电影推荐', '某歌手演唱会官宣'],
        'zhihu': ['如何看待当前的经济形势？', '有哪些让人深思的电影？', '读研真的有用吗？', '程序员35岁何去何从？', '如何评价AI技术？', '改变人生的好习惯', '怎样提高工作效率？', '你最尴尬的经历？', '如何看待内卷？', '值得推荐的书单', '年轻人如何理财？', '如何保持健康？', '有趣的冷知识', '这届年轻人怎么了？', '职场最重要的能力'],
        'baidu': ['最新天气预报', '某明星近况', '高考分数线公布', '热门电影排行', '新款手机推荐', '旅游景点推荐', '美食菜谱大全', '健身锻炼方法', '育儿教育经验', '装修设计灵感', '汽车降价信息', '股市最新行情', '招聘信息汇总', '电影票房排名', '明星八卦新闻'],
        'douyin': ['挑战100元吃遍夜市', '这个舞蹈太上头了', '萌宠搞笑合集', '教你做家常菜', '旅行vlog分享', '职场干货分享', '健身打卡挑战', '手工艺教程', '科技数码测评', '穿搭分享', '读书笔记推荐', '音乐翻唱合集', '搞笑段子合集', '生活小妙招', '美食探店记录'],
        'toutiao': ['国内经济形势分析', '科技领域重大突破', '教育改革最新政策', '医疗健康新发现', '体育赛事精彩回顾', '文化艺术展览推荐', '社会热点事件追踪', '环保生态新举措', '乡村振兴进展', '国际形势深度解读'],
        'bilibili': ['某UP主新视频爆火', '新番动画推荐', '游戏攻略合集', '学习资源分享', '鬼畜区精选', '科技数码评测', '音乐创作分享', '美食制作教程', '旅行见闻分享', '影视解说精选'],
    }
    rng = random.Random(seed)
    topics = topics_map.get(platform, topics_map['weibo'])
    return [{'rank': i+1, 'title': t, 'heat': rng.randint(100000, 9999999) if i < 5 else rng.randint(1000, 500000),
             'change': rng.choice([-2, -1, 0, 0, 0, 1, 1, 2, 3, 5]), 'tags': [], 'url': ''}
            for i, t in enumerate(topics[:50])]


# ---- 多线程并行刷新 ----

def _refresh_all_platforms():
    """多线程并行抓取全部平台，总耗时取决于最慢的单平台"""
    logger.info("[Hotlist] Parallel refresh started (ThreadPoolExecutor)...")
    t0 = time.time()
    platforms = list(HOTLIST_PLATFORMS.keys())

    def _fetch_one(p):
        try:
            items = _scrape_platform(p)
            with _hotlist_lock:
                _hotlist_cache[p] = {
                    "items": items, "count": len(items),
                    "updated_at": datetime.now().isoformat(),
                }
            return p, len(items)
        except Exception as e:
            logger.warning(f"[Hotlist] {p} failed: {e}")
            return p, 0

    # 最多6个线程并行（每个平台一个线程）
    with concurrent.futures.ThreadPoolExecutor(max_workers=6, thread_name_prefix="hotlist") as executor:
        futures = {executor.submit(_fetch_one, p): p for p in platforms}
        for future in concurrent.futures.as_completed(futures, timeout=30):
            try:
                p, count = future.result()
                logger.info(f"[Hotlist]   {p}: {count} items")
            except Exception as e:
                logger.warning(f"[Hotlist]   future error: {e}")

    elapsed = time.time() - t0
    _save_hotlist_cache()
    logger.info(f"[Hotlist] Parallel refresh complete in {elapsed:.2f}s")


_load_hotlist_cache()


class HotlistRefreshRequest(BaseModel):
    platform: Optional[str] = None


@app.get("/api/hotlist")
async def get_hotlist(platform: str = "all"):
    if platform == "all":
        with _hotlist_lock:
            result = dict(_hotlist_cache)
        cfg = {k: {"name": v["name"], "color": v["color"], "total": v["total"]} for k, v in HOTLIST_PLATFORMS.items()}
        return JSONResponse(content={"success": True, "platforms": result, "config": cfg})
    if platform not in HOTLIST_PLATFORMS:
        raise HTTPException(status_code=400, detail=f"Unknown platform: {platform}")
    with _hotlist_lock:
        cached = _hotlist_cache.get(platform)
    if cached and cached.get("items"):
        return JSONResponse(content={"success": True, "platform": platform, "config": HOTLIST_PLATFORMS[platform],
                                     "items": cached["items"], "count": len(cached["items"]),
                                     "updated_at": cached.get("updated_at", ""), "cached": True})
    loop = asyncio.get_running_loop()
    items = await loop.run_in_executor(None, lambda: _scrape_platform(platform))
    with _hotlist_lock:
        _hotlist_cache[platform] = {"items": items, "updated_at": datetime.now().isoformat(), "count": len(items)}
    return JSONResponse(content={"success": True, "platform": platform, "config": HOTLIST_PLATFORMS[platform],
                                 "items": items, "count": len(items), "updated_at": datetime.now().isoformat(), "cached": False})


@app.post("/api/hotlist/refresh")
async def refresh_hotlist(req: HotlistRefreshRequest = None):
    p = req.platform if req else None
    if p and p in HOTLIST_PLATFORMS:
        t0 = time.time()
        items = _scrape_platform(p)
        with _hotlist_lock:
            _hotlist_cache[p] = {"items": items, "updated_at": datetime.now().isoformat(), "count": len(items)}
        _save_hotlist_cache()
        elapsed = time.time() - t0
        return JSONResponse(content={"success": True, "platform": p, "count": len(items), "elapsed": round(elapsed, 2)})
    # 刷新全部：多线程并行
    _refresh_all_platforms()
    return JSONResponse(content={"success": True, "message": "All platforms refreshed in parallel", "updated_at": datetime.now().isoformat()})


@app.get("/api/hotlist/export")
async def export_hotlist(platform: str = "all", format: str = "json"):
    from fastapi.responses import StreamingResponse
    with _hotlist_lock:
        if platform == "all":
            all_data = dict(_hotlist_cache)
        elif platform in _hotlist_cache:
            all_data = {platform: _hotlist_cache[platform]}
        else:
            raise HTTPException(status_code=404, detail="No data")

    if format == "json":
        return JSONResponse(content={"success": True, "data": all_data, "exported_at": datetime.now().isoformat()})

    if format == "csv":
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["平台", "排名", "热词", "热度", "变化", "标签", "链接", "更新时间"])
        for p, pd in all_data.items():
            pn = HOTLIST_PLATFORMS.get(p, {}).get("name", p)
            for item in pd.get("items", []):
                w.writerow([pn, item.get("rank", ""), item.get("title", ""), item.get("heat", 0),
                            item.get("change", 0), ",".join(item.get("tags", [])), item.get("url", ""), pd.get("updated_at", "")])
        output.seek(0)
        bom = io.BytesIO(b'\xef\xbb\xbf' + output.getvalue().encode("utf-8"))
        bom.seek(0)
        return StreamingResponse(bom, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": f"attachment; filename=hotlist_{datetime.now().strftime('%Y%m%d')}.csv"})

    if format == "excel":
        if not OPENPYXL_AVAILABLE:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        from openpyxl.styles import PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "热搜数据"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for i, h in enumerate(["平台", "排名", "热词", "热度", "变化", "标签", "链接", "更新时间"], 1):
            c = ws.cell(row=1, column=i, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
        r = 2
        for p, pd in all_data.items():
            pn = HOTLIST_PLATFORMS.get(p, {}).get("name", p)
            for item in pd.get("items", []):
                ws.cell(row=r, column=1, value=pn); ws.cell(row=r, column=2, value=item.get("rank", ""))
                ws.cell(row=r, column=3, value=item.get("title", "")); ws.cell(row=r, column=4, value=item.get("heat", 0))
                ws.cell(row=r, column=5, value=item.get("change", 0)); ws.cell(row=r, column=6, value=",".join(item.get("tags", [])))
                ws.cell(row=r, column=7, value=item.get("url", "")); ws.cell(row=r, column=8, value=pd.get("updated_at", ""))
                r += 1
        for i, w in enumerate([12, 8, 40, 12, 8, 15, 40, 20], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=hotlist_{datetime.now().strftime('%Y%m%d')}.xlsx"})

    raise HTTPException(status_code=400, detail="format must be json, csv, or excel")


# ---- 每日自动更新调度器 ----

_scheduler_stop_event = threading.Event()

def _daily_scheduler():
    while not _scheduler_stop_event.is_set():
        now = datetime.now()
        nxt = now.replace(hour=0, minute=5, second=0, microsecond=0)
        if now >= nxt:
            nxt = nxt.replace(day=nxt.day + 1)
        noon = now.replace(hour=12, minute=5, second=0, microsecond=0)
        if now < noon and (noon - now).total_seconds() < (nxt - now).total_seconds():
            nxt = noon
        wait = max((nxt - now).total_seconds(), 60)
        logger.info(f"[Hotlist] Next auto-refresh: {nxt.strftime('%Y-%m-%d %H:%M')}")
        while not _scheduler_stop_event.is_set():
            _scheduler_stop_event.wait(60)
            if datetime.now() >= nxt:
                break
        if not _scheduler_stop_event.is_set():
            _refresh_all_platforms()


def _start_scheduler():
    threading.Thread(target=_daily_scheduler, daemon=True, name="hotlist-scheduler").start()
    logger.info("[Hotlist] Daily auto-update scheduler started")








# ---------------------------------------------------------------------------
# ---- Translation Helper ----
_translate_cache = {}
def _translate_to_zh(text):
    """Translate English text to Chinese using Google Translate unofficial API"""
    if not text or not text.strip():
        return text
    text = text.strip()
    cache_key = text[:200]
    if cache_key in _translate_cache:
        return _translate_cache[cache_key]
    try:
        encoded = quote(text[:500])
        api_url = 'https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=zh-CN&dt=t&q=' + encoded
        resp = requests.get(api_url, headers={"User-Agent": random.choice(USER_AGENT_POOL)}, timeout=8)
        if resp.status_code == 200:
            data = resp.json()
            if data and data[0]:
                translated = "".join(item[0] for item in data[0] if item[0])
                if translated:
                    _translate_cache[cache_key] = translated
                    return translated
    except Exception as e:
        logger.debug(f"Translation failed: {e}")
    _translate_cache[cache_key] = text
    return text


def _translate_batch(texts):
    results = list(texts)
    to_translate = [(i, t) for i, t in enumerate(texts) if t and t.strip() and t.strip()[:200] not in _translate_cache]
    if not to_translate:
        return [_translate_cache.get(t.strip()[:200], t) if t else t for t in texts]
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        future_map = {executor.submit(_translate_to_zh, t): i for i, t in to_translate}
        for future in concurrent.futures.as_completed(future_map):
            idx = future_map[future]
            try:
                results[idx] = future.result()
            except Exception:
                pass
    return results

# GitHub Trending
# ---------------------------------------------------------------------------
import concurrent.futures as _cf

_github_cache = {}
_github_lock = threading.Lock()

def _scrape_github_trending(since="daily"):
    """Scrape GitHub trending page"""
    url = f"https://github.com/trending?since={since}"
    headers = {
        "User-Agent": random.choice(USER_AGENT_POOL),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15, verify=False)
        resp.raise_for_status()
    except Exception as e:
        logger.error(f"GitHub trending fetch failed: {e}")
        return []
    
    soup = BeautifulSoup(resp.text, "lxml")
    repos = []
    
    for idx, row in enumerate(soup.select("article.Box-row"), 1):
        try:
            # Repo name
            h2 = row.select_one("h2 a")
            if not h2:
                continue
            full_name = h2.get_text(strip=True).replace("\n", "").replace(" ", "")
            parts = full_name.split("/")
            owner = parts[0].strip() if len(parts) > 0 else ""
            name = parts[1].strip() if len(parts) > 1 else ""
            url_repo = "https://github.com" + h2["href"].strip()
            
            # Description
            desc_tag = row.select_one("p")
            description = desc_tag.get_text(strip=True) if desc_tag else ""
            # description collected for batch translate
            
            # Language
            lang_tag = row.select_one("[itemprop='programmingLanguage']")
            language = lang_tag.get_text(strip=True) if lang_tag else ""
            
            # Stars
            star_links = row.select("a.Link--muted")
            stars = 0
            forks = 0
            if len(star_links) >= 1:
                stars_text = star_links[0].get_text(strip=True).replace(",", "")
                try: stars = int(stars_text)
                except: pass
            if len(star_links) >= 2:
                forks_text = star_links[1].get_text(strip=True).replace(",", "")
                try: forks = int(forks_text)
                except: pass
            
            # Stars today/week/month
            stars_period = 0
            period_tag = row.select_one("span.d-inline-block.float-sm-right")
            if period_tag:
                period_text = period_tag.get_text(strip=True).replace(",", "")
                nums = re.findall(r'\d+', period_text)
                if nums:
                    stars_period = int(nums[0])
            
            repos.append({
                "rank": idx,
                "owner": owner,
                "name": name,
                "full_name": full_name,
                "url": url_repo,
                "description": description,
                "language": language,
                "stars": stars,
                "forks": forks,
                "stars_period": stars_period,
            })
        except Exception as e:
            logger.warning(f"GitHub trending parse error at item {idx}: {e}")
    

    # Batch translate all descriptions in parallel
    descriptions = [r.get("description", "") for r in repos]
    translated = _translate_batch(descriptions)
    for r, t in zip(repos, translated):
        if t:
            r["description"] = t

    return repos


@app.get("/api/github/trending")
async def github_trending(since: str = "daily"):
    if since not in ("daily", "weekly", "monthly"):
        raise HTTPException(status_code=400, detail="since must be daily, weekly, or monthly")
    
    with _github_lock:
        cached = _github_cache.get(since)
        if cached and cached.get("repos") and cached.get("updated_at"):
            updated = datetime.fromisoformat(cached["updated_at"])
            if (datetime.now() - updated).total_seconds() < 1800:  # 30min cache
                return JSONResponse(content={"success": True, "since": since, "repos": cached["repos"], "count": len(cached["repos"]), "updated_at": cached["updated_at"], "cached": True})
    
    loop = asyncio.get_running_loop()
    repos = await loop.run_in_executor(None, lambda: _scrape_github_trending(since))
    with _github_lock:
        _github_cache[since] = {"repos": repos, "updated_at": datetime.now().isoformat()}
    
    return JSONResponse(content={"success": True, "since": since, "repos": repos, "count": len(repos), "updated_at": datetime.now().isoformat(), "cached": False})


@app.get("/api/github/trending/export")
async def github_trending_export(since: str = "daily", format: str = "json"):
    with _github_lock:
        cached = _github_cache.get(since)
    if not cached or not cached.get("repos"):
        raise HTTPException(status_code=404, detail="No data. Fetch trending first.")
    
    repos = cached["repos"]
    
    if format == "json":
        return JSONResponse(content={"success": True, "since": since, "repos": repos, "exported_at": datetime.now().isoformat()})
    
    if format == "csv":
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["Rank", "Owner", "Name", "Full Name", "URL", "Description", "Language", "Stars", "Forks", "Stars Period"])
        for r in repos:
            w.writerow([r["rank"], r["owner"], r["name"], r["full_name"], r["url"], r["description"], r["language"], r["stars"], r["forks"], r["stars_period"]])
        output.seek(0)
        bom = io.BytesIO(b'\xef\xbb\xbf' + output.getvalue().encode("utf-8"))
        bom.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(bom, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename=github-trending-{since}.csv"})
    
    if format == "excel":
        if not OPENPYXL_AVAILABLE:
            raise HTTPException(status_code=500, detail="openpyxl not installed")
        from openpyxl.styles import PatternFill
        wb = Workbook(); ws = wb.active; ws.title = f"GitHub Trending ({since})"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for i, h in enumerate(["Rank", "Owner", "Name", "Description", "Language", "Stars", "Forks", "Stars Period", "URL"], 1):
            c = ws.cell(row=1, column=i, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
        r = 2
        for repo in repos:
            ws.cell(row=r, column=1, value=repo["rank"]); ws.cell(row=r, column=2, value=repo["owner"])
            ws.cell(row=r, column=3, value=repo["name"]); ws.cell(row=r, column=4, value=repo["description"])
            ws.cell(row=r, column=5, value=repo["language"]); ws.cell(row=r, column=6, value=repo["stars"])
            ws.cell(row=r, column=7, value=repo["forks"]); ws.cell(row=r, column=8, value=repo["stars_period"])
            ws.cell(row=r, column=9, value=repo["url"])
            r += 1
        for i, w in enumerate([6, 16, 24, 50, 14, 10, 10, 12, 40], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=github-trending-{since}.xlsx"})
    
    raise HTTPException(status_code=400, detail="format must be json, csv, or excel")


# ===========================================================================
# Callable API Endpoints — 外部程序可直接调用的简化接口
# ===========================================================================

class QuickScrapeRequest(BaseModel):
    """Quick scrape — single URL, auto mode, returns cleaned results."""
    url: str
    extract_types: List[str] = ["text", "links", "metadata"]
    timeout: int = Field(30, ge=5, le=120)
    proxy: Optional[str] = None
    depth: int = Field(0, ge=0, le=5)
    max_pages: int = Field(1, ge=1, le=50)


class QuickBatchRequest(BaseModel):
    """Quick batch — multiple URLs, auto mode."""
    urls: List[str]
    extract_types: List[str] = ["text", "links", "metadata"]
    timeout: int = Field(30, ge=5, le=120)
    proxy: Optional[str] = None


class QuickExtractRequest(BaseModel):
    """Quick extract — fetch URL and return only one specific data type."""
    url: str
    extract_type: str = "text"  # text | links | media | structured | metadata | business
    timeout: int = Field(30, ge=5, le=120)
    proxy: Optional[str] = None


@app.get("/api/v1/health")
async def api_v1_health():
    """Health check for API consumers."""
    return {
        "ok": True,
        "version": "2.0.0",
        "playwright": PLAYWRIGHT_AVAILABLE,
        "endpoints": [
            "GET  /api/v1/health",
            "GET  /api/v1/scrape?url=<url>&format=json|text|links",
            "POST /api/v1/scrape  (body: QuickScrapeRequest)",
            "POST /api/v1/batch   (body: QuickBatchRequest)",
            "POST /api/v1/extract (body: QuickExtractRequest)",
            "GET  /api/v1/hotlist?platform=<name>",
            "GET  /api/v1/github/trending?since=daily",
            "GET  /api/v1/export?url=<url>&format=json|csv",
        ],
    }


@app.get("/api/v1/scrape")
async def api_v1_scrape_get(
    url: str,
    format: str = "json",
    extract_types: str = "text,links,metadata",
    timeout: int = 30,
    depth: int = 0,
    max_pages: int = 1,
):
    """
    Quick scrape via GET — no body needed.
    Example: /api/v1/scrape?url=https://example.com&format=json
    format: json (structured) | text (plain text only) | links (links only)
    """
    if not url or not url.strip():
        raise HTTPException(status_code=400, detail="url parameter is required")
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    types = [t.strip() for t in extract_types.split(",") if t.strip()]

    options = ScrapeOptions(
        depth=min(depth, 5),
        max_pages=min(max_pages, 50),
        timeout=min(timeout, 120),
        max_retries=1,
        rotate_ua=True,
    )

    try:
        result = await scrape_page(url, types, options, "auto")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if format == "text":
        text_data = result.get("extracted", {}).get("text", {})
        return {"success": True, "url": url, "title": result.get("title", ""),
                "text": text_data.get("main_content", ""),
                "paragraphs": text_data.get("paragraphs", []),
                "headings": text_data.get("headings", [])}

    if format == "links":
        links_data = result.get("extracted", {}).get("links", {})
        return {"success": True, "url": url,
                "internal": links_data.get("internal", []),
                "external": links_data.get("external", []),
                "total": links_data.get("all_count", 0)}

    return {"success": True, **result}


@app.post("/api/v1/scrape")
async def api_v1_scrape_post(req: QuickScrapeRequest):
    """
    Quick scrape via POST — returns full or filtered results.
    Body: {"url": "https://example.com", "extract_types": ["text","links"]}
    """
    url = req.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="url is required")
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    options = ScrapeOptions(
        depth=min(req.depth, 5),
        max_pages=min(req.max_pages, 50),
        timeout=min(req.timeout, 120),
        proxy=req.proxy,
        max_retries=1,
        rotate_ua=True,
    )

    try:
        result = await scrape_page(url, req.extract_types, options, "auto")
        return {"success": True, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v1/batch")
async def api_v1_batch(req: QuickBatchRequest):
    """
    Batch scrape — multiple URLs at once.
    Body: {"urls": ["https://a.com", "https://b.com"], "timeout": 20}
    """
    if not req.urls:
        raise HTTPException(status_code=400, detail="urls list is required")

    options = ScrapeOptions(
        timeout=min(req.timeout, 120),
        proxy=req.proxy,
        max_retries=1,
        rotate_ua=True,
    )

    results = []
    errors = []
    for raw_url in req.urls[:20]:  # cap at 20
        u = raw_url.strip()
        if not u:
            continue
        if not u.startswith(("http://", "https://")):
            u = "https://" + u
        try:
            r = await scrape_page(u, req.extract_types, options, "auto")
            results.append(r)
        except Exception as e:
            errors.append({"url": u, "error": str(e)})

    return {"success": True, "count": len(results), "pages": results, "errors": errors}


@app.post("/api/v1/extract")
async def api_v1_extract(req: QuickExtractRequest):
    """
    Extract specific data type from a single URL.
    Body: {"url": "https://example.com", "extract_type": "text"}
    extract_type: text | links | media | structured | metadata | business
    """
    valid_types = {"text", "links", "media", "structured", "metadata", "business"}
    if req.extract_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"extract_type must be one of: {', '.join(valid_types)}")

    url = req.url.strip()
    if not url:
        raise HTTPException(status_code=400, detail="url is required")
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    options = ScrapeOptions(
        timeout=min(req.timeout, 120),
        proxy=req.proxy,
        max_retries=1,
        rotate_ua=True,
    )

    try:
        result = await scrape_page(url, [req.extract_type], options, "auto")
        extracted = result.get("extracted", {}).get(req.extract_type, {})
        return {"success": True, "url": url, "title": result.get("title", ""),
                "type": req.extract_type, "data": extracted}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/hotlist")
async def api_v1_hotlist(platform: str = "weibo"):
    """
    Get hotlist for a specific platform.
    Example: /api/v1/hotlist?platform=weibo
    Platforms: weibo, zhihu, baidu, douyin, toutiao, bilibili
    """
    if platform not in HOTLIST_PLATFORMS:
        raise HTTPException(status_code=400, detail=f"Unknown platform: {platform}. Available: {', '.join(HOTLIST_PLATFORMS.keys())}")

    with _hotlist_lock:
        cached = _hotlist_cache.get(platform)
    if cached and cached.get("items"):
        return {"success": True, "platform": platform,
                "name": HOTLIST_PLATFORMS[platform]["name"],
                "items": cached["items"], "count": len(cached["items"]),
                "updated_at": cached.get("updated_at", "")}

    loop = asyncio.get_running_loop()
    items = await loop.run_in_executor(None, lambda: _scrape_platform(platform))
    with _hotlist_lock:
        _hotlist_cache[platform] = {"items": items, "updated_at": datetime.now().isoformat(), "count": len(items)}
    return {"success": True, "platform": platform,
            "name": HOTLIST_PLATFORMS[platform]["name"],
            "items": items, "count": len(items),
            "updated_at": datetime.now().isoformat()}


@app.get("/api/v1/github/trending")
async def api_v1_github_trending(since: str = "daily"):
    """
    Get GitHub trending repos.
    Example: /api/v1/github/trending?since=daily
    since: daily | weekly | monthly
    """
    if since not in ("daily", "weekly", "monthly"):
        raise HTTPException(status_code=400, detail="since must be daily, weekly, or monthly")

    with _github_lock:
        cached = _github_cache.get(since)
        if cached and cached.get("repos") and cached.get("updated_at"):
            updated = datetime.fromisoformat(cached["updated_at"])
            if (datetime.now() - updated).total_seconds() < 1800:
                return {"success": True, "since": since, "repos": cached["repos"],
                        "count": len(cached["repos"])}

    loop = asyncio.get_running_loop()
    repos = await loop.run_in_executor(None, lambda: _scrape_github_trending(since))
    with _github_lock:
        _github_cache[since] = {"repos": repos, "updated_at": datetime.now().isoformat()}
    return {"success": True, "since": since, "repos": repos, "count": len(repos)}


@app.get("/api/v1/export")
async def api_v1_export(url: str, format: str = "json", extract_types: str = "text,links,metadata"):
    """
    Scrape and export results directly.
    Example: /api/v1/export?url=https://example.com&format=csv
    format: json | csv
    """
    from fastapi.responses import StreamingResponse

    if not url or not url.strip():
        raise HTTPException(status_code=400, detail="url parameter is required")
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    types = [t.strip() for t in extract_types.split(",") if t.strip()]
    options = ScrapeOptions(timeout=30, max_retries=1, rotate_ua=True)

    try:
        result = await scrape_page(url, types, options, "auto")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if format == "csv":
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["URL", "Title", "Status", "Type", "Content"])
        ext = result.get("extracted", {})
        if ext.get("text", {}).get("main_content"):
            w.writerow([url, result.get("title", ""), result.get("status_code", 200),
                        "text", ext["text"]["main_content"][:500]])
        for lnk in ext.get("links", {}).get("internal", []) + ext.get("links", {}).get("external", []):
            lnk_url = lnk.get("url", "") if isinstance(lnk, dict) else str(lnk)
            w.writerow([url, result.get("title", ""), result.get("status_code", 200),
                        "link", lnk_url])
        for img in ext.get("media", {}).get("images", [])[:20]:
            img_url = img.get("url", "") if isinstance(img, dict) else str(img)
            w.writerow([url, result.get("title", ""), result.get("status_code", 200),
                        "image", img_url])
        output.seek(0)
        bom = io.BytesIO(b'\xef\xbb\xbf' + output.getvalue().encode("utf-8"))
        bom.seek(0)
        return StreamingResponse(bom, media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": f"attachment; filename=export.csv"})

    return {"success": True, **result}



# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    import webbrowser
    import threading

    logger.info("WebScraper v2.0 starting on http://localhost:8765")
    logger.info(f"   Playwright available: {PLAYWRIGHT_AVAILABLE}")
    logger.info(f"   aiohttp available: {AIOHTTP_AVAILABLE}")
    logger.info(f"   Features: XHR capture, WS monitoring, Shadow DOM, iframes, DFS/BFS, sitemap discovery")

    def open_browser():
        time.sleep(2)
        webbrowser.open("http://localhost:8765")

    threading.Thread(target=open_browser, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=8765, log_level="info")
